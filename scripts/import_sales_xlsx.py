import argparse
import hashlib
import json
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl


HEADER_ROW = 4
DATA_START_ROW = 5


def text(value: Any) -> str | None:
    if value is None:
        return None
    result = str(value).strip()
    return result or None


def number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return None


def parse_period(sheet_title: str, row_value: Any) -> tuple[str, str]:
    source = " ".join(part for part in [text(sheet_title), text(row_value)] if part)
    dates = []
    for token in source.replace(":", " ").split():
        try:
            dates.append(datetime.strptime(token, "%d.%m.%Y").date().isoformat())
        except ValueError:
            pass
    if len(dates) >= 2:
        return dates[0], dates[1]
    today = datetime.now().date().isoformat()
    return today, today


def stable_id(row: dict[str, Any]) -> str:
    payload = json.dumps(row, ensure_ascii=False, sort_keys=True)
    return hashlib.sha1(payload.encode("utf-8")).hexdigest()


def ensure_schema(connection: sqlite3.Connection) -> None:
    connection.executescript(
        """
        CREATE TABLE IF NOT EXISTS iiko_sales (
          id TEXT PRIMARY KEY,
          sale_date TEXT,
          dish_name TEXT NOT NULL,
          product_id TEXT,
          category TEXT,
          department TEXT,
          amount REAL NOT NULL DEFAULT 0,
          revenue REAL NOT NULL DEFAULT 0,
          raw_json TEXT NOT NULL,
          synced_at TEXT NOT NULL
        );
        """
    )
    columns = {row[1] for row in connection.execute("PRAGMA table_info(iiko_sales)")}
    additions = {
        "concept": "TEXT",
        "code": "TEXT",
        "group_name": "TEXT",
        "avg_price": "REAL",
        "avg_price_no_discount": "REAL",
        "revenue_no_discount": "REAL",
        "gross_profit": "REAL",
        "markup_percent": "REAL",
        "discount_sum": "REAL",
        "cost_per_unit": "REAL",
        "cost_total": "REAL",
        "cost_percent": "REAL",
    }
    for column, definition in additions.items():
        if column not in columns:
            connection.execute(f"ALTER TABLE iiko_sales ADD COLUMN {column} {definition}")
    connection.execute("CREATE INDEX IF NOT EXISTS idx_iiko_sales_date ON iiko_sales(sale_date)")
    connection.execute("CREATE INDEX IF NOT EXISTS idx_iiko_sales_dish ON iiko_sales(dish_name)")


def parse_workbook(path: Path) -> tuple[str, str, list[dict[str, Any]]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    period_from, period_to = parse_period(str(sheet.cell(1, 1).value), sheet.cell(2, 1).value)
    rows: list[dict[str, Any]] = []

    for row_index, row in enumerate(sheet.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW):
        name = text(row[1] if len(row) > 1 else None)
        if not name or name.lower().startswith("итого"):
            continue
        raw = {
            "sourceRow": row_index,
            "category": text(row[0] if len(row) > 0 else None) or "Без категории",
            "dishName": name,
            "avgPriceNoDiscount": number(row[2] if len(row) > 2 else None),
            "avgPrice": number(row[3] if len(row) > 3 else None),
            "amount": number(row[4] if len(row) > 4 else None) or 0,
            "revenueNoDiscount": number(row[5] if len(row) > 5 else None),
            "revenue": number(row[6] if len(row) > 6 else None) or 0,
            "grossProfit": number(row[8] if len(row) > 8 else None),
            "markupPercent": number(row[9] if len(row) > 9 else None),
            "grossProfitNoVat": number(row[10] if len(row) > 10 else None),
            "concept": text(row[11] if len(row) > 11 else None) or "Без концепции",
            "revenueShare": number(row[12] if len(row) > 12 else None),
            "code": text(row[13] if len(row) > 13 else None),
            "group": text(row[14] if len(row) > 14 else None),
            "date": text(row[15] if len(row) > 15 else None) or period_from,
            "fullName": text(row[16] if len(row) > 16 else None),
            "discountSum": number(row[17] if len(row) > 17 else None),
            "costPerUnit": number(row[18] if len(row) > 18 else None),
            "costTotal": number(row[19] if len(row) > 19 else None),
            "costPercent": number(row[20] if len(row) > 20 else None),
        }
        raw["id"] = stable_id(raw)
        rows.append(raw)

    return period_from, period_to, rows


def import_rows(db_path: Path, period_from: str, period_to: str, rows: list[dict[str, Any]], source: str) -> None:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    now = datetime.now().isoformat(timespec="seconds")
    with sqlite3.connect(db_path) as connection:
        ensure_schema(connection)
        connection.execute("DELETE FROM iiko_sales WHERE sale_date >= ? AND sale_date <= ?", (period_from, period_to))
        connection.executemany(
            """
            INSERT INTO iiko_sales (
              id, sale_date, dish_name, product_id, category, department, amount, revenue,
              concept, code, group_name, avg_price, avg_price_no_discount, revenue_no_discount,
              gross_profit, markup_percent, discount_sum, cost_per_unit, cost_total, cost_percent,
              raw_json, synced_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    row["id"],
                    row["date"],
                    row["dishName"],
                    row["code"],
                    row["category"],
                    row["concept"],
                    row["amount"],
                    row["revenue"],
                    row["concept"],
                    row["code"],
                    row["group"],
                    row["avgPrice"],
                    row["avgPriceNoDiscount"],
                    row["revenueNoDiscount"],
                    row["grossProfit"],
                    row["markupPercent"],
                    row["discountSum"],
                    row["costPerUnit"],
                    row["costTotal"],
                    row["costPercent"],
                    json.dumps({**row, "source": source}, ensure_ascii=False),
                    now,
                )
                for row in rows
            ],
        )


def main() -> None:
    parser = argparse.ArgumentParser(description="Import iiko sales Excel report into local SQLite.")
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("--db", type=Path, default=Path("data/iiko-chef.sqlite"))
    args = parser.parse_args()

    period_from, period_to, rows = parse_workbook(args.xlsx)
    import_rows(args.db, period_from, period_to, rows, f"excel:{args.xlsx.name}")
    revenue = sum(row["revenue"] for row in rows)
    amount = sum(row["amount"] for row in rows)
    concepts = sorted({row["concept"] for row in rows})
    print(json.dumps({"periodFrom": period_from, "periodTo": period_to, "rows": len(rows), "amount": amount, "revenue": revenue, "concepts": concepts}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
