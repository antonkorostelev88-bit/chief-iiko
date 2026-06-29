from __future__ import annotations

import argparse
import hashlib
import json
import re
import sqlite3
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl


KITCHEN_PLACES = {
    "Гриль/Гарниры",
    "Кондитерский цех",
    "Супы/Паста",
    "Япония",
    "Раздача",
    "Заготовочный",
    "Мясной",
    "Холодный цех",
    "Холодный цех - Закуски",
    "Бизнес-Ланч",
    "Бизнес-Ланч - Паста",
    "Овощной",
    "Завтраки",
    "Гриль/Гарниры - Супы/Паста",
    "Закуски",
    "Бизнез-Ланч - Супы",
    "Холодный цех(Море) - Кондитерский(ост)",
    "Гриль/Гарниры - Холодный цех",
    "Холодный цех - Кондитерский цех",
    "Холодный цех - Раздача",
    "Бизнес Ланч Супы/Паста - Гриль/Гарниры",
    "Япония - Гриль/Гарниры",
    "Холодный цех + Гриль/Гарниры (ДО ЭЛИС)",
}


@dataclass(frozen=True)
class RecipeRow:
    sheet: str
    parent_name: str
    ingredient_name: str
    quantity: float
    unit: str


@dataclass(frozen=True)
class ProductRow:
    id: str
    name: str
    kind: str
    article: str
    raw_json: str


@dataclass(frozen=True)
class ProductDefinition:
    name: str
    kind: str
    article: str
    code: str
    unit: str
    category: str
    group: str
    price: float | None
    raw: dict[str, Any]


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def normalize_key(value: Any) -> str:
    return normalize_text(value).casefold()


def normalize_article(value: Any) -> str:
    return re.sub(r"[^0-9A-Za-zА-Яа-я_-]+", "", str(value or "").replace("№", "").strip())


def normalize_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(" ", "").replace(",", "."))
    except ValueError:
        return None


def find_header_index(headers: list[str], *names: str) -> int:
    for name in names:
        if name in headers:
            return headers.index(name)
    raise RuntimeError("Не найдена колонка: " + " / ".join(names))


def parse_recipe_sheets(xlsx_path: Path) -> tuple[list[RecipeRow], dict[str, str], dict[str, tuple[float, str]]]:
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    rows: list[RecipeRow] = []
    parent_articles: dict[str, str] = {}
    parent_yields: dict[str, tuple[float, str]] = {}

    for sheet_name in ("ЛУБ", "ЛУПФ"):
        worksheet = workbook[sheet_name]
        header_values = [normalize_text(cell.value) for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
        parent_col = find_header_index(header_values, "Вхождение в")
        ingredient_col = find_header_index(header_values, "Состав", "Блюдо")
        quantity_col = find_header_index(header_values, "Кол-во")
        unit_col = find_header_index(header_values, "ед.из", "Ед.из")
        article_col = find_header_index(header_values, "Артикл IIKO", "Артикл")

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            parent_name = normalize_text(row[parent_col])
            ingredient_name = normalize_text(row[ingredient_col])
            if not parent_name or not ingredient_name:
                continue
            article = normalize_article(row[article_col])
            if article:
                parent_articles[normalize_key(parent_name)] = article
            quantity = normalize_number(row[quantity_col])
            unit = normalize_text(row[unit_col])
            if ingredient_name.casefold() == "итого":
                if quantity is not None:
                    parent_yields[normalize_key(parent_name)] = (quantity, unit)
                continue
            if normalize_key(ingredient_name) == normalize_key(parent_name):
                continue
            if quantity is None:
                continue
            rows.append(
                RecipeRow(
                    sheet=sheet_name,
                    parent_name=parent_name,
                    ingredient_name=ingredient_name,
                    quantity=quantity,
                    unit=unit,
                )
            )
    return rows, parent_articles, parent_yields


def stable_product_id(kind: str, name: str, article: str) -> str:
    source = "|".join(["approved", kind, normalize_key(name), normalize_article(article)])
    return "approved-" + hashlib.sha1(source.encode("utf-8")).hexdigest()[:24]


def parse_place(value: Any) -> str:
    text = normalize_text(value)
    if "Группа:" in text:
        return normalize_text(text.split("Группа:", 1)[0])
    return text


def parse_product_definitions(xlsx_path: Path) -> dict[str, ProductDefinition]:
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    definitions: dict[str, ProductDefinition] = {}

    goods_sheet = workbook["ЛУП"]
    goods_headers = [normalize_text(cell.value) for cell in next(goods_sheet.iter_rows(min_row=1, max_row=1))]
    goods_article_col = find_header_index(goods_headers, "Артикул iiko")
    goods_name_col = find_header_index(goods_headers, "Наименование в IIKO")
    goods_unit_col = find_header_index(goods_headers, "Ед. из.")
    goods_group_col = find_header_index(goods_headers, "Группа")
    for row in goods_sheet.iter_rows(min_row=2, values_only=True):
        name = normalize_text(row[goods_name_col])
        if not name:
            continue
        raw = {header or f"Колонка {index + 1}": value for index, (header, value) in enumerate(zip(goods_headers, row))}
        definitions.setdefault(
            normalize_key(name),
            ProductDefinition(
                name=name,
                kind="other",
                article=normalize_article(row[goods_article_col]),
                code="",
                unit=normalize_text(row[goods_unit_col]),
                category=normalize_text(row[goods_group_col]),
                group=normalize_text(row[goods_group_col]),
                price=None,
                raw={**raw, "source": "approved_goods"},
            ),
        )

    for sheet_name, kind in (("ЛУПФ", "semifinished"), ("ЛУБ", "dish")):
        worksheet = workbook[sheet_name]
        headers = [normalize_text(cell.value) for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
        article_col = find_header_index(headers, "Артикл IIKO", "Артикл")
        name_col = find_header_index(headers, "Вхождение в")
        category_col = find_header_index(headers, "Список классификации п/ф") if sheet_name == "ЛУПФ" else find_header_index(headers, "Раздел")
        place_col = find_header_index(headers, "Цех2", "Цех") if sheet_name == "ЛУПФ" else find_header_index(headers, "Место приготовления")
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            article = normalize_article(row[article_col])
            name = normalize_text(row[name_col])
            if not article or not name:
                continue
            raw = {header or f"Колонка {index + 1}": value for index, (header, value) in enumerate(zip(headers, row))}
            raw["Тип места приготовления"] = parse_place(row[place_col])
            raw["source"] = "approved_semifinished" if kind == "semifinished" else "approved_dish"
            definitions.setdefault(
                normalize_key(name),
                ProductDefinition(
                    name=name,
                    kind=kind,
                    article=article,
                    code="",
                    unit="кг" if kind == "semifinished" else "порц",
                    category=normalize_text(row[category_col]),
                    group=normalize_text(row[category_col]),
                    price=None,
                    raw=raw,
                ),
            )
    return definitions


def read_products(connection: sqlite3.Connection) -> list[ProductRow]:
    return [
        ProductRow(
            id=str(row["id"]),
            name=str(row["name"]),
            kind=str(row["kind"]),
            article=str(row["article"] or ""),
            raw_json=str(row["raw_json"] or "{}"),
        )
        for row in connection.execute("SELECT id, name, kind, article, raw_json FROM iiko_products")
    ]


def build_product_indexes(products: list[ProductRow]):
    by_name: dict[str, list[ProductRow]] = defaultdict(list)
    by_article: dict[str, list[ProductRow]] = defaultdict(list)
    for product in products:
        by_name[normalize_key(product.name)].append(product)
        article = normalize_article(product.article)
        if article:
            by_article[article].append(product)
    return by_name, by_article


def choose_product(candidates: list[ProductRow], preferred_kinds: tuple[str, ...]) -> ProductRow | None:
    if not candidates:
        return None
    for kind in preferred_kinds:
        for product in candidates:
            if product.kind == kind:
                return product
    return candidates[0]


def match_parent(
    parent_name: str,
    parent_articles: dict[str, str],
    by_name: dict[str, list[ProductRow]],
    by_article: dict[str, list[ProductRow]],
    preferred_kind: str,
) -> ProductRow | None:
    key = normalize_key(parent_name)
    article = parent_articles.get(key)
    if article:
        exact_article_candidates = [item for item in by_article.get(article, []) if normalize_key(item.name) == key]
        chosen = choose_product(exact_article_candidates or by_article.get(article, []), (preferred_kind,))
        if chosen:
            return chosen
    return choose_product(by_name.get(key, []), (preferred_kind,))


def match_ingredient(
    name: str,
    by_name: dict[str, list[ProductRow]],
    by_article: dict[str, list[ProductRow]],
    definitions: dict[str, ProductDefinition],
) -> ProductRow | None:
    key = normalize_key(name)
    by_exact_name = choose_product(by_name.get(key, []), ("semifinished", "other", "dish"))
    if by_exact_name:
        return by_exact_name
    definition = definitions.get(key)
    if definition and definition.article:
        return choose_product(by_article.get(definition.article, []), ("semifinished", "other", "dish"))
    return None


def production_place(product: ProductRow) -> str:
    try:
        raw = json.loads(product.raw_json)
    except json.JSONDecodeError:
        raw = {}
    return normalize_text(raw.get("Тип места приготовления")) or "Неопределенные"


def insert_missing_products(
    connection: sqlite3.Connection,
    definitions: dict[str, ProductDefinition],
    needed_names: set[str],
    products: list[ProductRow],
) -> int:
    existing_names = {normalize_key(product.name) for product in products}
    existing_articles = {normalize_article(product.article) for product in products if product.article}
    inserted = 0
    now = datetime.now(timezone.utc).isoformat()
    insert = connection.execute
    for key in sorted(needed_names):
        if key in existing_names:
            continue
        definition = definitions.get(key)
        if not definition:
            continue
        if definition.article and definition.article in existing_articles:
            continue
        product_id = stable_product_id(definition.kind, definition.name, definition.article)
        insert(
            """
            INSERT OR IGNORE INTO iiko_products (
              id, name, kind, type, article, code, measure_unit, category, group_name,
              category_id, category_name, group_id, group_display_name, price, raw_json, synced_at, is_local
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                product_id,
                definition.name,
                definition.kind,
                "Заготовка" if definition.kind == "semifinished" else "Блюдо" if definition.kind == "dish" else "Товар",
                definition.article or None,
                definition.code or None,
                definition.unit or None,
                definition.category or None,
                definition.group or None,
                None,
                definition.category or None,
                None,
                definition.group or None,
                definition.price,
                json.dumps(definition.raw, ensure_ascii=False, default=str),
                now,
                0,
            ),
        )
        existing_names.add(key)
        if definition.article:
            existing_articles.add(definition.article)
        inserted += 1
    return inserted


def backup_database(connection: sqlite3.Connection, db_path: Path, label: str) -> Path:
    backup_path = db_path.with_name(f"iiko-chef.before-{label}.{datetime.now(timezone.utc).isoformat().replace(':', '-').replace('.', '-')}.sqlite")
    connection.execute("VACUUM INTO '" + str(backup_path).replace("'", "''") + "'")
    return backup_path


def import_recipes_and_cleanup(db_path: Path, xlsx_path: Path, delete_non_kitchen_dishes: bool) -> dict[str, Any]:
    recipe_rows, parent_articles, parent_yields = parse_recipe_sheets(xlsx_path)
    product_definitions = parse_product_definitions(xlsx_path)
    by_parent: dict[tuple[str, str], list[RecipeRow]] = defaultdict(list)
    for row in recipe_rows:
        preferred_parent_kind = "dish" if row.sheet == "ЛУБ" else "semifinished"
        by_parent[(row.parent_name, preferred_parent_kind)].append(row)

    connection = sqlite3.connect(db_path)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    backup_path = backup_database(connection, db_path, "approved-recipes-import")
    products = read_products(connection)
    now = datetime.now(timezone.utc).isoformat()
    needed_names = {normalize_key(row.parent_name) for row in recipe_rows} | {normalize_key(row.ingredient_name) for row in recipe_rows}

    matched_parents = 0
    inserted_rows = 0
    inserted_products = 0
    updated_yields = 0
    unmatched_parents: Counter[str] = Counter()
    unmatched_ingredients: Counter[str] = Counter()

    try:
        connection.execute("BEGIN")
        inserted_products = insert_missing_products(connection, product_definitions, needed_names, products)
        products = read_products(connection)
        by_name, by_article = build_product_indexes(products)
        for (parent_name, parent_kind), rows in by_parent.items():
            parent = match_parent(parent_name, parent_articles, by_name, by_article, parent_kind)
            if not parent:
                unmatched_parents[parent_name] += 1
                continue
            prepared_rows: list[tuple[str, str, float, str, int, str]] = []
            for sort_order, row in enumerate(rows):
                ingredient = match_ingredient(row.ingredient_name, by_name, by_article, product_definitions)
                if not ingredient:
                    unmatched_ingredients[row.ingredient_name] += 1
                    continue
                prepared_rows.append((parent.id, ingredient.id, row.quantity, row.unit, sort_order, now))
            if not prepared_rows:
                continue
            connection.execute("DELETE FROM local_recipe_items WHERE dish_product_id = ?", (parent.id,))
            connection.executemany(
                "INSERT INTO local_recipe_items (dish_product_id, ingredient_product_id, gross_quantity, unit, sort_order, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                prepared_rows,
            )
            yield_value = parent_yields.get(normalize_key(parent_name))
            if yield_value:
                connection.execute(
                    """
                    INSERT INTO product_production_settings (product_id, batch_volume, batch_unit, updated_at)
                    VALUES (?, ?, ?, ?)
                    ON CONFLICT(product_id) DO UPDATE SET
                      batch_volume = excluded.batch_volume,
                      batch_unit = excluded.batch_unit,
                      updated_at = excluded.updated_at
                    """,
                    (parent.id, yield_value[0], yield_value[1], now),
                )
                updated_yields += 1
            matched_parents += 1
            inserted_rows += len(prepared_rows)

        deleted_dishes = 0
        if delete_non_kitchen_dishes:
            products_after_import = read_products(connection)
            non_kitchen_dishes = [product for product in products_after_import if product.kind == "dish" and production_place(product) not in KITCHEN_PLACES]
            delete_recipe_rows = connection.execute
            for product in non_kitchen_dishes:
                delete_recipe_rows("DELETE FROM local_recipe_items WHERE dish_product_id = ? OR ingredient_product_id = ?", (product.id, product.id))
                delete_recipe_rows("DELETE FROM product_production_settings WHERE product_id = ?", (product.id,))
                delete_recipe_rows("DELETE FROM iiko_recipes WHERE product_id = ?", (product.id,))
                delete_recipe_rows("DELETE FROM iiko_products WHERE id = ?", (product.id,))
            deleted_dishes = len(non_kitchen_dishes)

        connection.execute(
            "INSERT INTO sync_runs (started_at, finished_at, source, status, total_products, total_recipes) VALUES (?, ?, ?, ?, ?, ?)",
            (now, datetime.now(timezone.utc).isoformat(), f"approved-recipes:{xlsx_path.name}", "success", 0, inserted_rows),
        )
        connection.execute("PRAGMA optimize")
        connection.commit()
        connection.execute("VACUUM")
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()

    return {
        "backup": str(backup_path),
        "excel_recipe_rows": len(recipe_rows),
        "excel_parents": len(by_parent),
        "matched_parents": matched_parents,
        "inserted_rows": inserted_rows,
        "inserted_missing_products": inserted_products,
        "updated_yields": updated_yields,
        "deleted_non_kitchen_dishes": deleted_dishes,
        "unmatched_parents": len(unmatched_parents),
        "unmatched_ingredients": len(unmatched_ingredients),
        "unmatched_parent_sample": [name for name, _ in unmatched_parents.most_common(20)],
        "unmatched_ingredient_sample": [name for name, _ in unmatched_ingredients.most_common(30)],
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Import approved dish and semifinished recipes from Excel into local SQLite.")
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("--db", type=Path, default=Path("data/iiko-chef.sqlite"))
    parser.add_argument("--keep-all-dishes", action="store_true")
    args = parser.parse_args()

    result = import_recipes_and_cleanup(args.db, args.xlsx, not args.keep_all_dishes)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
