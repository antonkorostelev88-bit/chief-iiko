from __future__ import annotations

import argparse
import hashlib
import json
import re
import sqlite3
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl

TYPE_MAP = {
    "Товар": "other",
    "Заготовка": "semifinished",
    "Блюдо": "dish",
    "Модификатор": "other",
    "Услуга": "other",
}
DEFAULT_NAME_LEVELS = 9
START_SEPARATORS = set(" ,;\n\r\t")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def normalize_price(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def stable_id(row_number: int, kind: str, name: str, article: str, code: str, group: str) -> str:
    source = "|".join([kind, article, code, name, group, str(row_number)])
    return "excel-" + hashlib.sha1(source.encode("utf-8")).hexdigest()[:24]


def product_name(row: tuple[Any, ...], name_level_count: int) -> str:
    for value in reversed(row[:name_level_count]):
        text = normalize_text(value)
        if text:
            return text
    return ""


def raw_json(headers: list[str], row: tuple[Any, ...], name_level_count: int) -> str:
    data: dict[str, Any] = {}
    for index, value in enumerate(row):
        if index < name_level_count:
            key = f"Название уровень {index + 1}"
        else:
            key = headers[index] if index < len(headers) and headers[index] else f"Колонка {index + 1}"
        if isinstance(value, datetime):
            data[key] = value.isoformat(sep=" ")
        else:
            data[key] = value
    data["source"] = "excel_nomenclature"
    return json.dumps(data, ensure_ascii=False, default=str)


def build_matcher(products: list[dict[str, Any]]):
    names = sorted({normalize_text(item["name"]) for item in products if item["kind"] in {"dish", "semifinished"}}, key=len, reverse=True)
    return names


def match_parent_names(text: str, candidate_names: list[str]) -> list[str]:
    source = normalize_text(text)
    result: list[str] = []
    position = 0
    while position < len(source):
        if source[position] in START_SEPARATORS:
            position += 1
            continue
        matched = ""
        for name in candidate_names:
            if not source.startswith(name, position):
                continue
            end = position + len(name)
            tail = source[end:].lstrip()
            # The item in "Включено в" must end here. This prevents short cards
            # like "Соус" from catching every longer "Соус ..." parent.
            if not tail or tail.startswith(",") or tail.startswith(";"):
                matched = name
                break
        if matched:
            result.append(matched)
            position += len(matched)
        else:
            position += 1
    return result


def find_header_row(worksheet) -> tuple[int, list[Any]]:
    for row_number, row in enumerate(worksheet.iter_rows(min_row=1, max_row=12, values_only=True), start=1):
        values = [normalize_text(value) for value in row]
        if "Название" in values and "Тип" in values:
            return row_number, list(row)
    raise RuntimeError("Не найдена строка заголовков с колонками Название и Тип")


def parse_workbook(xlsx_path: Path):
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header_row_number, headers = find_header_row(worksheet)
    header_index = {name: index for index, name in enumerate(headers) if name}

    required = ["Название", "Артикул", "Код", "Ед. измерения", "Цена, р.", "Категория", "Включено в", "Тип", "Группа"]
    missing = [name for name in required if name not in header_index]
    if missing:
        raise RuntimeError("В Excel не найдены колонки: " + ", ".join(missing))

    name_level_count = max(1, header_index.get("Артикул", DEFAULT_NAME_LEVELS))
    products: list[dict[str, Any]] = []
    included_by_child: dict[str, str] = {}
    type_counts = Counter()

    for row_number, row in enumerate(worksheet.iter_rows(min_row=header_row_number + 1, values_only=True), start=header_row_number + 1):
        excel_type = normalize_text(row[header_index["Тип"]])
        type_counts[excel_type or "Пусто"] += 1
        if excel_type not in TYPE_MAP:
            continue
        name = product_name(row, name_level_count)
        if not name:
            continue
        kind = TYPE_MAP[excel_type]
        article = normalize_text(row[header_index["Артикул"]])
        code = normalize_text(row[header_index["Код"]])
        unit = normalize_text(row[header_index["Ед. измерения"]])
        category = normalize_text(row[header_index["Категория"]])
        group = normalize_text(row[header_index["Группа"]])
        product_id = stable_id(row_number, kind, name, article, code, group)
        product = {
            "id": product_id,
            "row_number": row_number,
            "name": name,
            "kind": kind,
            "type": excel_type,
            "article": article or None,
            "code": code or None,
            "measure_unit": unit or None,
            "category": category or None,
            "group": group or None,
            "price": normalize_price(row[header_index["Цена, р."]]),
            "raw_json": raw_json(headers, row, name_level_count),
        }
        products.append(product)
        included = normalize_text(row[header_index["Включено в"]])
        if included:
            included_by_child[product_id] = included

    name_candidates = build_matcher(products)
    products_by_name = defaultdict(list)
    for product in products:
        products_by_name[normalize_text(product["name"])].append(product)

    recipe_edges: list[tuple[str, str, str | None]] = []
    unmatched = []
    edge_seen = set()
    for child in products:
        included = included_by_child.get(child["id"])
        if not included:
            continue
        parent_names = match_parent_names(included, name_candidates)
        if not parent_names:
            unmatched.append({"child": child["name"], "included": included[:500]})
            continue
        for parent_name in parent_names:
            parents = products_by_name.get(parent_name, [])
            if not parents:
                continue
            for parent in parents:
                if parent["id"] == child["id"] or parent["kind"] not in {"dish", "semifinished"}:
                    continue
                key = (parent["id"], child["id"])
                if key in edge_seen:
                    continue
                edge_seen.add(key)
                recipe_edges.append((parent["id"], child["id"], child["measure_unit"]))

    return products, recipe_edges, type_counts, unmatched

def import_to_sqlite(db_path: Path, products: list[dict[str, Any]], recipe_edges: list[tuple[str, str, str | None]], source: str):
    db_path.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(db_path)
    connection.execute("PRAGMA foreign_keys = ON")
    now = datetime.now(timezone.utc).isoformat()
    try:
        connection.execute("BEGIN")
        connection.execute("DELETE FROM local_recipe_items")
        connection.execute("DELETE FROM iiko_recipes")
        connection.execute("DELETE FROM product_production_settings")
        connection.execute("DELETE FROM iiko_products")
        connection.execute("DELETE FROM iiko_references")
        connection.execute("DELETE FROM iiko_raw_payloads")
        connection.executemany(
            """
            INSERT INTO iiko_products (
              id, name, kind, type, article, code, measure_unit, category, group_name,
              category_id, category_name, group_id, group_display_name, price, raw_json, synced_at, is_local
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    item["id"],
                    item["name"],
                    item["kind"],
                    item["type"],
                    item["article"],
                    item["code"],
                    item["measure_unit"],
                    item["category"],
                    item["group"],
                    None,
                    item["category"],
                    None,
                    item["group"],
                    item["price"],
                    item["raw_json"],
                    now,
                    0,
                )
                for item in products
            ],
        )
        connection.executemany(
            "INSERT INTO local_recipe_items (dish_product_id, ingredient_product_id, gross_quantity, unit, sort_order, created_at) VALUES (?, ?, ?, ?, ?, ?)",
            [(parent_id, child_id, None, unit, index, now) for index, (parent_id, child_id, unit) in enumerate(recipe_edges)],
        )
        connection.execute(
            "INSERT INTO sync_runs (started_at, finished_at, source, status, total_products, total_recipes) VALUES (?, ?, ?, ?, ?, ?)",
            (now, now, source, "success", len(products), len(recipe_edges)),
        )
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


def main():
    parser = argparse.ArgumentParser(description="Import iiko nomenclature Excel export into local SQLite.")
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("--db", type=Path, default=Path("data/iiko-chef.sqlite"))
    args = parser.parse_args()

    products, recipe_edges, type_counts, unmatched = parse_workbook(args.xlsx)
    import_to_sqlite(args.db, products, recipe_edges, f"excel:{args.xlsx.name}")

    kind_counts = Counter(product["kind"] for product in products)
    categories = {product["category"] for product in products if product["category"]}
    recipe_parents = {parent_id for parent_id, _, _ in recipe_edges}
    print(json.dumps({
        "products": len(products),
        "goods": kind_counts["other"],
        "semifinished": kind_counts["semifinished"],
        "dishes": kind_counts["dish"],
        "categories": len(categories),
        "recipeRows": len(recipe_edges),
        "recipeParents": len(recipe_parents),
        "unmatchedIncludedRows": len(unmatched),
        "typeCountsInExcel": type_counts,
        "unmatchedSamples": unmatched[:20],
    }, ensure_ascii=False, default=str, indent=2))


if __name__ == "__main__":
    main()



